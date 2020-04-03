import pyautogui
from openpyxl import Workbook
from openpyxl import load_workbook
from pynput.mouse import Listener
from pynput.mouse import Events
from pynput import mouse

pyautogui.FAILSAFE = True

caminho = '/home/birohp/Documentos/pybots/relatorio.xlsx'
arquivo_excel = load_workbook(caminho)
wb = arquivo_excel.active

#pyautogui.click(1088,41)

def on_click(x, y, button, pressed):
    if pressed:
        print ('({0}, {1})'.format(x, y, button))

# The event listener will be running in this block
with mouse.Events() as events:
    # Block at most one second
    event = events.get(1.0)
    if event is None:
        print('You did not interact with the mouse within one second')
    else:
        print('Received event {}'.format(event))

'''
max_linha = wb.max_row
max_coluna = wb.max_column
for i in range(1, max_linha + 1):
    pyautogui.doubleClick(1082,340)
    for j in range(1, max_coluna + 1):
        pyautogui.typewrite((wb.cell(row=i, column=j).value))
        pyautogui.PAUSE = 1
        pyautogui.press('tab')

pyautogui.press('tab')
pyautogui.PAUSE = 1
'''