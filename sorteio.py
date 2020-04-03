import pyautogui as gui
import random
import sys
from time import sleep
from openpyxl import Workbook
from openpyxl import load_workbook

gui.FAILSAFE = True

path = '/home/birohp/Documentos/bots/sorteio.xlsx'
f = load_workbook(path)
wb = f.active
qt_names = int(sys.argv[1])
timer = sys.argv[2]
max_line = wb.max_row

line = ''
j = 1
x = 1
pos = gui.position()

if qt_names < 1:
    print('First parameter must be greater than 1')
    sys.exit()

if timer != 'y' and timer != 'n':
    print('Second parameter must be \'y\' or \'n\'')
    sys.exit()

for i in range(1, max_line + 1):
    line = line + wb.cell(row=i, column=1).value + ' '
    j = j + 1
    if j != qt_names + 1:
        continue
    gui.click(pos)
    sleep(2) if timer == 'y' else sleep(0)
    gui.typewrite(line)
    sleep(2) if timer == 'y' else sleep(0)
    gui.press('tab')
    sleep(2) if timer == 'y' else sleep(0)
    gui.press('enter')
    r = random.randint(20, 60)
    sleep(r) if timer == 'y' else sleep(0)
    print(str(x) + ': ' + str(r) + 's')
    x = x + 1
    line = ''
    j = 1

print('Done: ' + str(x-1) + ' comments!')
